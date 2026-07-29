from datetime import datetime
from io import BytesIO

import os
import uuid

from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    send_file
)

from openpyxl import Workbook
from openpyxl.styles import Font

from services.pdf_service import PDFService

from database import db
from models.produtos import Produto
from models.categoria import Categoria
from models.fornecedor import Fornecedor


produtos_bp = Blueprint(
    "produtos",
    __name__,
    url_prefix="/produtos"
)


UPLOAD_FOLDER = "static/uploads"

EXTENSOES_PERMITIDAS = {
    "png",
    "jpg",
    "jpeg",
    "webp"
}


def imagem_permitida(nome):

    return (
        "." in nome and
        nome.rsplit(".", 1)[1].lower()
        in EXTENSOES_PERMITIDAS
    )



def salvar_imagem(arquivo):

    if not arquivo or arquivo.filename == "":
        return None


    if not imagem_permitida(arquivo.filename):
        return None


    extensao = arquivo.filename.rsplit(".", 1)[1].lower()


    nome = (
        str(uuid.uuid4())
        +
        "."
        +
        extensao
    )


    os.makedirs(
        UPLOAD_FOLDER,
        exist_ok=True
    )


    caminho = os.path.join(
        UPLOAD_FOLDER,
        nome
    )


    arquivo.save(caminho)


    return nome



# ==========================
# LISTAR
# ==========================

@produtos_bp.route("/")
def listar():

    produtos = Produto.query.order_by(
        Produto.nome
    ).all()


    return render_template(
        "produtos/listar.html",
        produtos=produtos
    )



# ==========================
# NOVO
# ==========================

@produtos_bp.route(
    "/novo",
    methods=["GET", "POST"]
)
def novo():

    categorias = Categoria.query.order_by(
        Categoria.nome
    ).all()


    fornecedores = Fornecedor.query.order_by(
        Fornecedor.nome
    ).all()



    if request.method == "POST":


        imagem = salvar_imagem(
            request.files.get("imagem")
        )


        produto = Produto(

            codigo=request.form["codigo"],

            nome=request.form["nome"],

            descricao=request.form["descricao"],

            preco=float(
                request.form["preco"]
            ),

            quantidade=int(
                request.form["quantidade"]
            ),

            estoque_minimo=int(
                request.form["estoque_minimo"]
            ),

            categoria_id=int(
                request.form["categoria_id"]
            ),

            fornecedor_id=int(
                request.form["fornecedor_id"]
            ),

            imagem=imagem

        )


        db.session.add(produto)

        db.session.commit()



        flash(
            "Produto cadastrado com sucesso!",
            "success"
        )


        return redirect(
            url_for("produtos.listar")
        )



    return render_template(
        "produtos/cadastrar.html",
        categorias=categorias,
        fornecedores=fornecedores
    )



# ==========================
# EDITAR
# ==========================

@produtos_bp.route(
    "/editar/<int:id>",
    methods=["GET", "POST"]
)
def editar(id):

    produto = Produto.query.get_or_404(id)


    categorias = Categoria.query.order_by(
        Categoria.nome
    ).all()


    fornecedores = Fornecedor.query.order_by(
        Fornecedor.nome
    ).all()



    if request.method == "POST":


        produto.codigo = request.form["codigo"]

        produto.nome = request.form["nome"]

        produto.descricao = request.form["descricao"]

        produto.preco = float(
            request.form["preco"]
        )

        produto.quantidade = int(
            request.form["quantidade"]
        )

        produto.estoque_minimo = int(
            request.form["estoque_minimo"]
        )

        produto.categoria_id = int(
            request.form["categoria_id"]
        )

        produto.fornecedor_id = int(
            request.form["fornecedor_id"]
        )



        nova_imagem = salvar_imagem(
            request.files.get("imagem")
        )


        if nova_imagem:

            produto.imagem = nova_imagem



        db.session.commit()



        flash(
            "Produto atualizado!",
            "success"
        )


        return redirect(
            url_for("produtos.listar")
        )



    return render_template(
        "produtos/editar.html",
        produto=produto,
        categorias=categorias,
        fornecedores=fornecedores
    )



# ==========================
# EXCLUIR
# ==========================

@produtos_bp.route("/excluir/<int:id>")
def excluir(id):

    produto = Produto.query.get_or_404(id)


    db.session.delete(produto)

    db.session.commit()


    flash(
        "Produto removido!",
        "warning"
    )


    return redirect(
        url_for("produtos.listar")
    )



# ==========================
# EXCEL
# ==========================

@produtos_bp.route("/exportar-excel")
def exportar_excel():

    produtos = Produto.query.order_by(
        Produto.nome
    ).all()


    wb = Workbook()

    ws = wb.active

    ws.title = "Produtos"


    cabecalho = [
        "Código",
        "Nome",
        "Categoria",
        "Fornecedor",
        "Quantidade",
        "Preço",
        "Estoque Mínimo"
    ]


    for coluna, valor in enumerate(cabecalho, start=1):

        celula = ws.cell(
            row=1,
            column=coluna
        )

        celula.value = valor

        celula.font = Font(
            bold=True
        )


    linha = 2


    for produto in produtos:

        ws.cell(linha,1).value = produto.codigo
        ws.cell(linha,2).value = produto.nome
        ws.cell(linha,3).value = produto.categoria.nome
        ws.cell(linha,4).value = produto.fornecedor.nome
        ws.cell(linha,5).value = produto.quantidade
        ws.cell(linha,6).value = produto.preco
        ws.cell(linha,7).value = produto.estoque_minimo

        linha += 1


    arquivo = BytesIO()

    wb.save(arquivo)

    arquivo.seek(0)


    return send_file(
        arquivo,
        as_attachment=True,
        download_name=f"produtos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



# ==========================
# PDF
# ==========================

@produtos_bp.route("/exportar-pdf")
def exportar_pdf():

    produtos = Produto.query.order_by(
        Produto.nome
    ).all()


    pdf = PDFService.gerar_produtos(
        produtos
    )


    return send_file(
        pdf,
        as_attachment=True,
        download_name=f"produtos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
        mimetype="application/pdf"
    )